"""
Data Processing Utilities

This script contains functions to process, clean, and analyze datasets used in the research project.
Key functionalities include:
- Data import from various formats (.csv, .xls, .xlsx).
- Data cleaning and dictionary creation.
- Outlier detection and removal.
- Handling missing data with mapping or imputation.
- Generating data dictionaries for better data understanding.
- Support for interactive mapping and user-defined data transformations.

Dependencies:
- pandas
- numpy
- tqdm
- openpyxl
- sklearn
- statsmodels

Usage:
1. Import the functions into your main script or notebook.
2. Use the `clean` function for full data pipeline processing.
3. Customize outlier detection, mapping, or data validation as needed.

Authors: P. Sitthirat et al
Version: 1.0
License: MIT License
"""

# Standard library imports
import os
import re

# Third-party library imports
import pandas as pd
import numpy as np
from tqdm import tqdm
from tabulate import tabulate
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
import statsmodels.api as sm
from IPython.display import clear_output

class DataHandler:
    """
    A class for handling data import and dictionary creation.
    - Importing data
    - Creating the dictionary
    -
    """
    
    @staticmethod
    def data_import(file_name, data_dir):
        """
        Import a dataset from the specified directory.

        Parameters:
        - file_name (str): The name of the file (without extension).
        - data_dir (str): The directory where the data file is stored.

        Returns:
        - tuple:
        - pd.DataFrame: The imported dataset as a DataFrame.
        - str: Path to the created data dictionary.

        Raises:
            FileNotFoundError: If no file with the specified name is found.
            ValueError: If the file format is unsupported.
        """
        
        # Determine the file path and extension
        for ext in ['.csv', '.xlsx', '.xls']:
            file_path = os.path.join(data_dir, f'{file_name}{ext}')
            if os.path.exists(file_path):
                break
        else:
            raise FileNotFoundError(f"No file found for {file_name} with supported extensions (.csv, .xlsx, .xls)")

        print(f"Importing data: {file_name}{ext} ---")
        
        # Load dataset based on file type
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, low_memory=False)
        elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            while True:
                sheet_name = input("Please enter the sheet name: ")
                if sheet_name is None:
                    sheet_name = 'Sheet1'
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    break
                except ValueError:
                    print("Sheet name does not exist. Please try again.")
        else:
            raise ValueError("Unsupported file format")
        
        print(f"Imported {file_name}{ext} successfully")
        dict_path = DataHandler.create_dictionary(df, file_name, data_dir)

        return df, dict_path

    @staticmethod
    def create_dictionary(df, file_name, data_dir):
        """
        Creates a data dictionary from a DataFrame.

        Parameters:
        - df (pd.DataFrame): The DataFrame for which the dictionary is created.
        - file_name (str): The base file name for saving the dictionary.
        - data_dir (str): The directory to save the dictionary.

        Returns:
        - str: Path to the saved data dictionary.
        """
        
        data_dict = {
            'Include': [],
            'Index': [],
            'Source Name': [],
            'Column Name': [],
            'Column Label': [],
            'Data Type': [],
            'Required': [],
            'Unique': [],
            'Range': [],
            'Min': [],
            'Max': [],
            'Regex Pattern': [],
            'Default Value': [],
            'Null Value': [],
            'Format': [],
            'Examples': []
        }
        
        for index, col in tqdm(enumerate(df.columns), total=len(df.columns), desc="Creating dictionary"):
            data_dict['Index'].append(index+1)
            data_dict['Source Name'].append(col)
            data_dict['Data Type'].append(str(df[col].dtype))
            data_dict['Required'].append(not df[col].isnull().any())
            data_dict['Unique'].append(df[col].nunique() == len(df[col]))
            data_dict['Examples'].append(df[col].dropna().unique()[:5])
            
            # Determine the range for numerical and categorical data
            if pd.api.types.is_numeric_dtype(df[col]):
                data_dict['Range'].append((df[col].min(), df[col].max()))
                data_dict['Min'].append(df[col].min())
                data_dict['Max'].append(df[col].max())
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                data_dict['Range'].append((df[col].min(), df[col].max()))
                data_dict['Min'].append(df[col].min())
                data_dict['Max'].append(df[col].max())
            else:
                unique_vals = sorted(map(str, df[col].dropna().unique()))
                if len(unique_vals) <= 10:  # Arbitrarily choosing 10 as a cutoff for display
                    data_dict['Range'].append(f"Categorical with {len(unique_vals)} unique values: \n {unique_vals}")
                    data_dict['Min'].append(None)
                    data_dict['Max'].append(None)
                else:
                    data_dict['Range'].append(f"Categorical with {len(unique_vals)} unique values: \n {unique_vals[:5]} etc.")
                    data_dict['Min'].append(None)
                    data_dict['Max'].append(None)
            
            # Append None for other properties
            data_dict['Include'].append(None)
            data_dict['Column Name'].append(None)
            data_dict['Column Label'].append(None)
            data_dict['Regex Pattern'].append(None)
            data_dict['Default Value'].append(None)
            data_dict['Null Value'].append(None)
            data_dict['Format'].append(None)

        # Convert the dictionary to a DataFrame
        dict = pd.DataFrame(data_dict)

        # Define the path for the Excel file
        excel_path = f"{data_dir}/data_dict-{file_name}.xlsx"
        sheet_name = f"Dict-{file_name}_update"
        
        # Open the existing workbook using openpyxl
        if os.path.exists(excel_path):
            book = load_workbook(excel_path)
            # Remove the existing sheet if it exists
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            # Save the changes
            book.save(excel_path)
            book.close()
        
        # Check if the file exists, and then use ExcelWriter with mode='a' to append
        # If the file does not exist, mode='w' creates a new file
        if os.path.exists(excel_path):
            mode = 'a'
        else:
            mode = 'w'

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode) as writer:
            # Write the DataFrame to a specific sheet
            dict.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Created data dictionary successfully: saved to {excel_path}")

        return excel_path

class DataCleaning:
    """
    A class for cleaning and preprocessing data.
    - Updating column names and types
    - Validating data types
    - Removing duplications
    - Full cleaning pipeline
    """

    @staticmethod
    def update_column_name(df, validate_dict_path):
        """
        Updates column names in the DataFrame based on a validation dictionary.

        Parameters:
        - df (pd.DataFrame): The DataFrame to update.
        - validate_dict_path (str): Path to the validation dictionary.

        Returns:
        - pd.DataFrame: DataFrame with updated column names.
        
        Raises:
            ValueError: If the sheet name 'validate' does not exist in the validation dictionary file.
        """
        
        # Load the necessary columns from the validate sheet
        try:
            validate_dict = pd.read_excel(validate_dict_path, sheet_name="validate", usecols=["Source Name", "Column Name"])
        except ValueError:
            raise ValueError("The sheet name 'validate' does not exist in the provided validation dictionary file.")

        # Create a dictionary from the DataFrame for renaming
        rename_dict = dict(zip(validate_dict['Source Name'], validate_dict['Column Name']))

        # Filter out keys that are not in the DataFrame's columns
        rename_dict = {k: v for k, v in rename_dict.items() if k in df.columns}

        # Rename columns in one go
        df = df.rename(columns=rename_dict)

        print("Updated column names successfully")

        return df
    
    @staticmethod
    def validate_input(input_value, data_type, format):
        """
        Validate and convert the user's input to the expected data type and format. 
        Supports handling null values explicitly as 'none', 'null', or an empty string.

        Parameters:
        - input_value (str): The input value to validate and convert.
        - data_type (str): The target data type for validation. 
                        Options: 'str', 'object', 'date', 'datetime', 'time', 'int64', 'float64'.
        - format (str, optional): The expected format for date/time conversion (e.g., '%Y-%m-%d %H:%M:%S').

        Returns:
        - Union[str, int, float, datetime, None]: The validated and converted value.
                                                  Returns 'null' for explicit nulls.
                                                  Returns None for invalid conversions.

        Raises:
            ValueError: If `data_type` is invalid or unsupported.
        """
        
        input_value = input_value.strip()
        if input_value.lower() in ['none', 'null', '']:
            return 'null'  # Special marker for intentional nulls

        try:
            if data_type == 'str' or data_type == 'object':
                return str(input_value)
            elif data_type in ['date', 'datetime', 'time']:
                return pd.to_datetime(input_value, format=format)
            elif data_type in ['int64', 'float64']:
                return pd.to_numeric(input_value)
        except ValueError:
            return None  # None still represents a conversion error

    @staticmethod
    def update_column_type(df, dict_dir, file_name, output_dir, column_spc=None):
        """
        Updates column data types in the DataFrame based on a validation dictionary.
        Handles error correction and supports time-to-datetime conversion.

        Parameters:
        - df (pd.DataFrame): The DataFrame to update.
        - dict_dir (str): Directory containing the data dictionary.
        - file_name (str): Base file name for referencing validation and output files.
        - output_dir (str): Directory to save output and error files.
        - column_spc (str, optional): Specific column to update (default is None for all columns).

        Returns:
        - tuple:
          - pd.DataFrame: Updated DataFrame.
          - str: Path to the cleaned or error-marked data file.

        Raises:
            FileNotFoundError: If the validation dictionary file is not found.
        """
        
        # Import validate dict
        validate_dict_path = f'{dict_dir}/data_dict-{file_name}.xlsx'
        validate_dict = pd.read_excel(validate_dict_path, sheet_name="validate")
        
        # Create dataframe for manipulate
        df_mod = df.copy()
        i = 0
        e = 0
        total = len(validate_dict)
        
        if column_spc:
            validate_dict = validate_dict[validate_dict['Column Name'] == column_spc]

        error_dir = os.path.join(output_dir, f'error-{file_name}')
        if not os.path.exists(error_dir):
            os.makedirs(error_dir)
        else:
            error_files = sorted([f for f in os.listdir(error_dir) if f.endswith('.csv')])
            if len(error_files) > 0:
                print(f"Found {len(error_files)} files to process for error correction before updating column type.")
                df_mod = DataManipulation.replace_error(df_mod, error_dir, file_name)

        for _, row in validate_dict.iterrows():
            i += 1
            column = row['Column Name']
            new_type = row['Data Type']
            format = row.get('Format', None)
            
            if column in df.columns:
            
                # Cleaning and preparing null values list
                if pd.notna(row['Null Value']):
                    null_values = [x.replace('\xa0', ' ') for x in row['Null Value'].split(',')]
                    null_values = [s.replace('"', '') for s in null_values]
                else:
                    null_values = []  # Empty list if 'Null Value' is NaN
                    
                # Replace specified null values in the DataFrame
                for null_value in null_values:
                    print(f"Replace {null_value} in {column}")
                    df_mod[column] = df_mod[column].replace({null_value: None})  # Stripping spaces
                
                # For time-only data, aading the date to them
                if new_type == 'time':
                    date_column = None
                    while date_column == None:
                        date_column = input(f"Found the time data in {column}, please identify the date column for this column: ")
                        if date_column in df_mod.columns:
                            df_mod = DataManipulation.add_date(df_mod, column, date_column)
                        else:
                            print(f"Column {date_column} not found in DataFrame.")
                            date_column = None
                
                errors = []
                
                try:
                    if new_type in ['date', 'datetime', 'time']:
                        df_mod[column] = pd.to_datetime(df_mod[column], format=format)
                    elif new_type in ['float64']:
                        df_mod[column] = pd.to_numeric(df_mod[column])
                    elif new_type == 'str':
                        df_mod[column] = df_mod[column].astype('string')
                    print(f"Successfully converted {column} to {new_type}")
                except:
                    for idx, value in tqdm(df_mod[column].items(), total=df_mod[column].shape[0], desc=f"Failed converted {column} to {new_type} \nFinding the errors: "):
                        try:
                            if pd.isna(value):
                                continue  # Skip conversion for NaN values
                            if new_type == 'str' or new_type == 'object':
                                df_mod.at[idx, column] = str(value)
                            elif new_type in ['date', 'datetime', 'time']:
                                df_mod.at[idx, column] = pd.to_datetime(value, format=format)
                            elif new_type in ['int64', 'float64']:
                                df_mod.at[idx, column] = pd.to_numeric(value)
                        except (ValueError, TypeError):
                            errors.append({'Error': value})
        
            if len(errors) > 0:
                e += 1   
                error_df = pd.DataFrame(errors)
                
                error_path = os.path.join(error_dir, f"{e:02}_{file_name}_{column}-error.csv")
                error_df.to_csv(error_path, index=False)
                print(f"Found {len(error_df)} errors in {column}: saved to {error_path}")
                    
        if e >= 1:
            print(f"There are errors in {e} columns and need to be modified")
            clean_path = f"Need error modification at {error_dir}"
        else:
            print(f"There is no errors in any columns. Saving the cleaned file --")           
            clean_path = f'{output_dir}/{file_name}_cleaned.csv'
            df_mod.to_csv(clean_path, index=False)
            print(f"Updated column data successfully without errors: processed data file were exported to {clean_path}")
            DataHandler.create_dictionary(df_mod, file_name, dict_dir)
        
        return df_mod, clean_path

    @staticmethod
    def remove_duplicates(df, subset=None, keep='first'):
        """
        Remove duplicate rows from a DataFrame.

        Parameters:
        - df (pd.DataFrame): The DataFrame from which duplicates will be removed.
        - subset (list, optional): Columns to consider for identifying duplicates. Defaults to None, meaning all columns.
        - keep (str, optional): Determines which duplicates (if any) to keep.
            - 'first': (default) Keep the first occurrence and drop subsequent duplicates.
            - 'last': Keep the last occurrence and drop earlier duplicates.
            - False: Drop all duplicates.

        Returns:
        - pd.DataFrame: A DataFrame with duplicates removed.

        Prints:
            A message indicating the number of duplicates removed.
        """
        
        duplicates = df.duplicated(keep=keep)
        duplicate_count = duplicates.sum()
        
        print(f"Removing {duplicate_count} duplicates")
        
        df_unique = df.drop_duplicates(subset=subset, keep=keep)
        
        print(f"Removed {duplicate_count} duplicates successfully")
        
        return df_unique

    @staticmethod
    def clean(file, data_dir, output_dir, clear=True):
        """
        Perform a full cleaning pipeline for the given dataset, including:
        - Importing the dataset.
        - Updating column names based on the validation dictionary.
        - Updating column types and handling errors.
        - Removing duplicate rows.

        Parameters:
        - file (str): Name of the file (without extension) to process.
        - data_dir (str): Directory where the raw data file is located.
        - output_dir (str): Directory to save the cleaned dataset and error files.
        - clear (bool, optional): Whether to clear the console output after processing. Defaults to True.

        Returns:
        - tuple:
          - pd.DataFrame: Cleaned DataFrame.
          - pd.DataFrame: Raw DataFrame before cleaning.
          - str: Path to the validation dictionary used.
        """
        
        df_raw, dict_path = DataHandler.data_import(file, data_dir)
        df = DataCleaning.update_column_name(df_raw, dict_path)
        df, clean_path = DataCleaning.update_column_type(df, data_dir, file, output_dir)
        df = DataCleaning.remove_duplicates(df, keep='last')
        
        if clear == True:
            clear_output()
        
        # Cleaning summary
        clean_summary = {
            'File name': file,
            'Path to data dictionary': dict_path,
            'Path to cleaned dataset': clean_path
        }

        # Printing out the summary
        for key, value in clean_summary.items():
            print(f"{key:25}: {value}")
        
        return df, df_raw, dict_path

class DataManipulation:
    """
    A class for manupulating the data
    - Adding date to time-only data.
    - String splitting based on patterns.
    - Replacing error values.
    - Handling outliers using various methods.
    - Mapping and filling missing values.
    """
    
    @staticmethod
    def add_date(df, time_column, date_column):
        """
        Adds date into the time variable before transforming them to date-time type.

        Parameters:
        - df (pd.DataFrame): The DataFrame to update.
        - time_column (str): The column containing time values.
        - date_column (str): The column containing date values.

        Returns:
        - pd.DataFrame: Updated DataFrame with combined date and time.
        """
        
        df[time_column] = df[date_column].astype(str) + ' ' + df[time_column]
        return df

    @staticmethod 
    def str_split(input_str, pattern, number_of_groups):
        """
        Split a string into groups based on a regex pattern.

        Parameters:
        - input_str (str): The input string to split.
        - pattern (str): The regex pattern to match the input string.
        - number_of_groups (int): The number of capturing groups in the regex pattern.

        Returns:
        - list: A list containing the matched groups if the pattern matches, otherwise a list of None values.

        Example:
            >>> DataManipulation.str_split("2023-12-14", r"(\d{4})-(\d{2})-(\d{2})", 3)
            ['2023', '12', '14']
        """
        
        if isinstance(input_str, str):
            match = re.match(pattern, input_str)
            if match:
                return [match.group(i) for i in range(1, number_of_groups + 1)]
        return [None] * number_of_groups

    @staticmethod
    def replace_error(df, error_dir, file_name):
        """
        Replaces erroneous values in the DataFrame based on error files.

        Parameters:
        - df (pd.DataFrame): The DataFrame to update.
        - error_dir (str): Directory containing error correction files.
        - file_name (str): Base file name to match error files.

        Returns:
        - pd.DataFrame: Updated DataFrame with errors replaced.
        """
        # List error files
        error_files = sorted([f for f in os.listdir(error_dir) if f.endswith('.csv')])
        
        for i, error_file in enumerate(tqdm(error_files, desc="Processing error files"), start=1):
            # Extract column name from the file name
            column_name = error_file.replace(f'{file_name}_', '').replace('-error.csv', '')
            
            if column_name in df.columns:
                # Load the error CSV
                error_df = pd.read_csv(os.path.join(error_dir, error_file))
                
                # Create an array of 'error' where 'edit' is null
                null_values = set(error_df[error_df['Edit'].isna()]['Error'].values)
                mask = df[column_name].isin(null_values)
                df.loc[mask, column_name] = None
                
                # Create a mapping from error dictionary values
                error_dict = error_df.dropna(subset=['Edit']).set_index('Error')['Edit'].to_dict()
                
                # Replace values in the DataFrame
                df[column_name] = df[column_name].apply(lambda x: error_dict.get(x, x))
                
            else:
                print(f"Column {column_name} not found in DataFrame.")

        return df

    @staticmethod
    def outlier_percentile(df, parameter, percentile=90, direction='more than', impute=None, print_outliers=False):
        """
        Detect and handle outliers in a column using percentile-based thresholds.

        This method identifies outliers by calculating a percentile threshold and comparing values in the specified column against this threshold. Outliers are flagged and can optionally be imputed or replaced with NaN based on the chosen handling method.

        Parameters:
        - df (pd.DataFrame): The input DataFrame containing the column to analyze.
        - parameter (str): The column name to check for outliers.
        - percentile (int, optional): Percentile value to use for threshold calculation. Default is 90.
        - direction (str, optional): The direction to identify outliers:
          - 'more than': Values greater than the threshold.
          - 'less than': Values less than the threshold.
          - 'more than and equal to': Values greater than or equal to the threshold.
          - 'less than and equal to': Values less than or equal to the threshold.
        - impute (str, optional): Specifies how to handle flagged outliers. Options:
          - 'forward': Forward-fill missing values after flagging outliers.
          - 'backward': Backward-fill missing values after flagging outliers.
          - None: Leave flagged outliers as NaN (default).
        - print_outliers (bool, optional): If True, prints detected outliers for review. Default is False.

        Returns:
        - tuple:
          - pd.DataFrame: Updated DataFrame with outliers flagged and optionally imputed.
          - bool: A flag indicating whether any outliers were detected and handled.

        Raises:
            ValueError: If an invalid `direction` or `impute` method is provided.

        Notes:
            - This method requires a non-empty column with valid numeric data for percentile calculation.
            - Outliers are determined based on the specified direction and percentile threshold.
            - The `impute` parameter provides flexibility in how flagged outliers are handled, allowing users to maintain data continuity.

        Example:
            >>> updated_df, has_outliers = DataManipulation.outlier_percentile(
                    df=data,
                    parameter='column_name',
                    percentile=95,
                    direction='more than',
                    impute=None,
                    print_outliers=True
                )
        """
         
        # Filter valid data
        columns_to_check = [parameter]
        if not isinstance(columns_to_check, list):
            raise ValueError("Parameter must be a list of column names.")
        valid_data = df[parameter].dropna()
        if valid_data.empty:
            print(f"No valid values found for parameter '{parameter}'.")
            return df

        # Calculate the percentile threshold
        threshold = np.percentile(valid_data[parameter].values, percentile)

        modified = False  # Initialize modified flag

        # Identify outliers based on the specified direction
        if direction == 'more than':
            valid_data['outlier'] = valid_data.apply(lambda row: '*' if row[parameter] > threshold else '', axis=1)
        elif direction == 'less than':
            valid_data['outlier'] = valid_data.apply(lambda row: '*' if row[parameter] < threshold else '', axis=1)
        elif direction == 'more than and equal to':
            valid_data['outlier'] = valid_data.apply(lambda row: '*' if row[parameter] >= threshold else '', axis=1)
        elif direction == 'less than and equal to':
            valid_data['outlier'] = valid_data.apply(lambda row: '*' if row[parameter] <= threshold else '', axis=1)
        else:
            raise ValueError("Direction must be 'more than', 'less than', 'more than and equal to', or 'less than and equal to'.")

        # Extract outliers
        outliers = valid_data[valid_data['outlier'] == '*']  
        
        # Optionally print outliers
        if not outliers.empty:
            modified = True  # Mark as modified
            if print_outliers:
                print(f"Outliers detected in '{parameter}' based on {percentile}th percentile ({direction}):")
                print(tabulate(valid_data, headers='keys'))

        # Handle outliers
        if valid_data['outlier'].any():
            if impute == 'forward':
                valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                valid_data[parameter] = valid_data[parameter].ffill()
            elif impute == 'backward':
                valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                valid_data[parameter] = valid_data[parameter].bfill()
            elif impute is None:
                valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
            else:
                raise ValueError("Invalid impute method. Use 'forward', 'backward', or None.")
            modified = True

            if print_outliers:
                print(tabulate(valid_data, headers='keys'))

        # Replace outliers in the original DataFrame
        valid_data = valid_data.set_index('index')
        df.update(valid_data)
        
        return df, modified

    @staticmethod
    def outlier_iqr(df, parameter, type='mild', impute=None, print_outliers=False):
        """
        Detect and handle outliers in a column using the Interquartile Range (IQR) method.

        This method identifies outliers based on the specified threshold (1.5 times IQR for mild outliers and 3 times IQR for extreme outliers). Outliers are optionally flagged and handled through imputation or replacement with NaN.

        Parameters:
        - df (pd.DataFrame): The input DataFrame containing the column to analyze.
        - parameter (str): The column name in which to detect and handle outliers.
        - type (str, optional): Type of outlier detection:
          - 'mild': Flags outliers beyond 1.5 times the IQR (default).
          - 'extreme': Flags outliers beyond 3 times the IQR.
        - impute (str, optional): Specifies how to handle flagged outliers. Options:
          - 'forward': Forward-fill missing values after flagging outliers.
          - 'backward': Backward-fill missing values after flagging outliers.
          - None: Leave flagged outliers as NaN (default).
        - print_outliers (bool, optional): If True, prints detected outliers for review. Default is False.

        Returns:
        - tuple:
          - pd.DataFrame: Updated DataFrame with outliers replaced or handled.
          - bool: A flag indicating whether any outliers were detected and handled.

        Raises:
            ValueError: If the `type` parameter is invalid or if there are insufficient data points for meaningful IQR analysis.

        Notes:
            - This method requires at least four valid observations for IQR calculation.
            - For columns with zero IQR, the method falls back to robust alternatives such as the Median Absolute Deviation (MAD) or mean ± standard deviation.
            - The `impute` parameter provides flexibility in how flagged outliers are handled, ensuring data continuity.

        Example:
            >>> updated_df, has_outliers = DataManipulation.outlier_iqr(
                    df=data,
                    parameter='column_name',
                    type='mild',
                    impute=None,
                    print_outliers=True
                )
        """
        
        # Filter valid data
        columns_to_check = [parameter]
        if not isinstance(columns_to_check, list):
            raise ValueError("Parameter must be a list of column names.")
        valid_data = df[parameter].dropna()
        if valid_data.empty:
            print(f"No valid values found for parameter '{parameter}'.")
            return df
        
        if type == 'mild':
            threshold = 1.5
        elif type == 'extreme':
            threshold = 3
        else:
            raise ValueError("Invalid 'type'. Use 'mild' or 'extreme'.")
        
        modified = False  # Initialize modified flag

        if len(valid_data[parameter].values) >= 4:  # At least four values are needed for meaningful IQR calculation
            # Calculate IQR
            p25 = np.percentile(valid_data[parameter].values, 25)
            p75 = np.percentile(valid_data[parameter].values, 75)
            iqr = p75 - p25

            if iqr > 0:  # Regular IQR method
                threshold_max = p75 + (threshold * iqr)
                threshold_min = p25 - (threshold * iqr)
            else:  # Fallback to a robust measure: Median Absolute Deviation (MAD)
                median = np.median(valid_data[parameter])
                mad = np.median(np.abs(valid_data[parameter].values - median))

                if mad > 0:  # Use MAD method
                    threshold_max = median + (threshold * mad * 1.4826)  # Scale MAD
                    threshold_min = median - (threshold * mad * 1.4826)
                else:  # Fallback to mean ± std dev
                    mean = np.mean(valid_data[parameter].values)
                    std_dev = np.std(valid_data[parameter].values)
                    threshold_max = mean + (threshold * std_dev)
                    threshold_min = mean - (threshold * std_dev)

            valid_data['outlier'] = valid_data.apply(lambda row: '*' if (row[parameter] > threshold_max) or (row[parameter] < threshold_min) else '', axis=1)
            outliers = valid_data[valid_data['outlier'] == '*'] 

            # Optionally print outliers
            if not outliers.empty:
                modified = True  # Mark as modified
                if print_outliers:
                    print(f"Outliers detected in '{parameter}' based on {threshold} times of interquartile range (IQR):")
                    print(tabulate(valid_data, headers='keys'))

            # Handle outliers
            if valid_data['outlier'].any():
                if impute == 'forward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].ffill()
                elif impute == 'backward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].bfill()
                elif impute is None:
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                else:
                    raise ValueError("Invalid impute method. Use 'forward', 'backward', or None.")
                modified = True

                if print_outliers:
                    print(tabulate(valid_data, headers='keys'))

            # Replace outliers in the original DataFrame
            valid_data = valid_data.set_index('index')
            df.update(valid_data)
        
        return df, modified

    @staticmethod
    def outlier_moving(df, parameter, window=3, method='mean', threshold=2, impute=None, print_outliers=False):
        """
        Detect and handle outliers using a moving window-based method.

        This method identifies outliers by calculating a rolling statistic (mean or median) within a moving window and comparing values against dynamically calculated bounds. The bounds are determined using the rolling standard deviation or Median Absolute Deviation (MAD), scaled by a specified threshold.

        Parameters:
        - df (pd.DataFrame): The input DataFrame containing the column to analyze.
        - parameter (str): The column name to check for outliers.
        - window (int, optional): The number of subdivisions for the moving window. Default is 3.
        - method (str, optional): Method to calculate the rolling statistic:
          - 'mean': Uses the rolling mean and standard deviation.
          - 'median': Uses the rolling median and MAD. Default is 'mean'.
        - threshold (float, optional): Multiplier for the rolling standard deviation or MAD to determine outliers. Default is 2.
        - impute (str, optional): Specifies how to handle flagged outliers. Options:
          - 'forward': Forward-fill missing values after flagging outliers.
          - 'backward': Backward-fill missing values after flagging outliers.
          - None: Leave flagged outliers as NaN (default).
        - print_outliers (bool, optional): If True, prints detected outliers for review. Default is False.

        Returns:
        - tuple:
          - pd.DataFrame: Updated DataFrame with outliers flagged and optionally imputed.
          - bool: A flag indicating whether any outliers were detected and handled.

        Raises:
            ValueError: If the `method` parameter is invalid or if the column contains insufficient data for analysis.

        Notes:
            - The window size is dynamically calculated based on the length of the data and the specified subdivisions.
            - Outliers are determined by comparing values to rolling upper and lower bounds.
            - The `impute` parameter provides flexibility in handling flagged outliers, allowing for forward-fill or backward-fill.

        Example:
            >>> updated_df, has_outliers = DataManipulation.outlier_moving(
                    df=data,
                    parameter='column_name',
                    window=5,
                    method='median',
                    threshold=3,
                    impute=None,
                    print_outliers=True
                )
        """
        
        # Filter valid data
        columns_to_check = [parameter]
        if not isinstance(columns_to_check, list):
            raise ValueError("Parameter must be a list of column names.")
        valid_data = df[parameter].dropna()
        if valid_data.empty:
            print(f"No valid values found for parameter '{parameter}'.")
            return df
        
        modified = False
        
        if len(valid_data[parameter].values) > window:
            window_size = int(np.ceil(len(valid_data[parameter].values) / window))
            if method == 'mean':
                # Calculate moving average and standard deviation
                moving_avg = valid_data[parameter].rolling(window=window_size, center=True).mean()
                moving_std = valid_data[parameter].rolling(window=window_size, center=True).std()
                valid_data['upper_bound'] = moving_avg + threshold * moving_std
                valid_data['lower_bound'] = moving_avg - threshold * moving_std
            elif method == 'median':
                # Calculate moving median and MAD
                moving_median = valid_data[parameter].rolling(window=window_size, center=True).median()
                mad = valid_data[parameter].rolling(window=window_size, center=True).apply(
                    lambda x: max(np.median(np.abs(x - np.median(x))), 1), raw=True)
                valid_data['mad'] = mad
                valid_data['upper_bound'] = moving_median + threshold * mad * 1.4826
                valid_data['lower_bound'] = moving_median - threshold * mad * 1.4826
            else:
                raise ValueError("Invalid method. Use 'mean' or 'median'.")    
            
            # Detect outliers
            valid_data['outlier'] = valid_data.apply(
                lambda row: '*' if row[parameter] > row['upper_bound'] or row[parameter] < row['lower_bound'] else '', axis=1
            )

            outliers = valid_data[valid_data['outlier'] == '*']        
            
            # Optionally print outliers
            if not outliers.empty:
                modified = True  # Mark as modified
                if print_outliers:
                    print("\nOutliers detected based on OLS leave-one-out regression:")
                    print(tabulate(valid_data, headers='keys'))

            # Handle outliers
            if valid_data['outlier'].any():
                if impute == 'forward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].ffill()
                elif impute == 'backward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].bfill()
                elif impute is None:
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                else:
                    raise ValueError("Invalid impute method. Use 'forward', 'backward', or None.")
                modified = True

                if print_outliers:
                    print(tabulate(valid_data, headers='keys'))

            # Replace outliers in the original DataFrame
            valid_data = valid_data.set_index('index')
            df.update(valid_data)

        return df, modified

    @staticmethod
    def outlier_regression(df, parameter, predictor, threshold=2, impute=None, print_outliers=False):
        """
        Detect and handle anomalies in a column using a regression model.

        This method identifies anomalies by fitting a linear regression model to the specified predictor and parameter columns. It calculates prediction bounds based on residual standard deviation, and flags data points falling outside these bounds as anomalies. Optionally, flagged anomalies can be imputed or replaced with NaN.

        Parameters:
        - df (pd.DataFrame): The input DataFrame containing the data to analyze.
        - parameter (str): The dependent variable (y) to assess for anomalies.
        - predictor (str): The independent variable (x) used for regression.
        - threshold (float, optional): Multiplier for the residual standard deviation to define anomaly bounds. Default is 2.
        - impute (str, optional): Specifies how to handle flagged anomalies. Options:
          - 'forward': Forward-fill missing values after flagging anomalies.
          - 'backward': Backward-fill missing values after flagging anomalies.
          - None: Leave flagged anomalies as NaN (default).
        - print_outliers (bool, optional): If True, prints detected anomalies for review. Default is False.

        Returns:
        - tuple:
          - pd.DataFrame: Updated DataFrame with anomalies flagged and optionally handled.
          - bool: A flag indicating whether any anomalies were detected and handled.

        Raises:
            ValueError: If parameter and predictor columns are invalid or missing.

        Notes:
            - This method requires at least two valid data points to perform regression.
            - Anomalies are flagged based on their residuals falling outside the upper and lower prediction bounds.
            - The `impute` parameter provides flexibility in handling flagged anomalies, allowing for forward-fill or backward-fill.

        Example:
            >>> updated_df, has_anomalies = DataManipulation.outlier_regression(
                    df=data,
                    parameter='dependent_var',
                    predictor='independent_var',
                    threshold=2.5,
                    impute=None,
                    print_outliers=True
                )
        """
        
        # Prepare data for regression
        columns_to_check = [parameter, predictor]
        if not isinstance(columns_to_check, list):
            raise ValueError("Parameter and predictor must be a list of column names.")
        valid_data = df[[parameter, predictor]].dropna()
        x = valid_data[predictor].values.reshape(-1, 1)
        y = valid_data[parameter].values
        
        modified = False  # Initialize modified flag

        if len(valid_data) > 1:
            
            # Fit regression model
            model = LinearRegression()
            model.fit(x, y)

            # Predict values
            predictions = model.predict(x)
            residuals = y - predictions
            residual_std = np.std(residuals)

            # Calculate bounds
            valid_data['predicted'] = predictions
            valid_data['upper_bound'] = predictions + threshold * residual_std
            valid_data['lower_bound'] = predictions - threshold * residual_std
            valid_data['residual'] = residuals
            valid_data['outlier'] = valid_data.apply(
                lambda row: '*' if row[parameter] > row['upper_bound'] or row[parameter] < row['lower_bound'] else '', axis=1
            )

            outliers = valid_data[valid_data['outlier'] == '*']        
            
            # Optionally print outliers
            if not outliers.empty:
                modified = True  # Mark as modified
                if print_outliers:
                    print("\nOutliers detected based on OLS leave-one-out regression:")
                    print(tabulate(valid_data, headers='keys'))

            # Handle outliers
            if valid_data['outlier'].any():
                if impute == 'forward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].ffill()
                elif impute == 'backward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].bfill()
                elif impute is None:
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                else:
                    raise ValueError("Invalid impute method. Use 'forward', 'backward', or None.")
                modified = True

                if print_outliers:
                    print(tabulate(valid_data, headers='keys'))

            # Replace outliers in the original DataFrame
            valid_data = valid_data.set_index('index')
            df.update(valid_data)

        return df, modified

    @staticmethod
    def outlier_loo(df, parameter, predictor, include_columns, mse_threshold_ratio=0.5, impute=None, print_outliers=False):
        """
        Detect and handle outliers by assessing their influence on OLS regression parameters using leave-one-out (LOO) analysis.

        This method identifies influential data points that significantly impact the regression model by removing one observation at a time and comparing the mean squared error (MSE) of the reduced models with the full model. Data points that fall below a specified threshold ratio are flagged as outliers.

        Parameters:
        - df (pd.DataFrame): The input DataFrame containing the data to analyze.
        - parameter (str): The dependent variable (y) to assess for outliers.
        - predictor (str): The independent variable (x) used for regression.
        - include_columns (list): Additional columns required for filtering valid data points.
        - mse_threshold_ratio (float, optional): The threshold ratio for MSE comparison. Points with MSE below this ratio of the full model's MSE are flagged as outliers. Default is 0.5.
        - impute (str, optional): Specifies how to handle flagged outliers. Options:
          - 'forward': Forward-fill missing values after flagging outliers.
          - 'backward': Backward-fill missing values after flagging outliers.
          - None: Leave flagged outliers as NaN. Default is None.
        - print_outliers (bool, optional): If True, prints the detected outliers. Default is False.

        Returns:
        - tuple:
          - pd.DataFrame: The updated DataFrame with outliers flagged and optionally imputed.
          - bool: A flag indicating whether any outliers were detected and handled.

        Raises:
            ValueError: If `include_columns` is not a list or any required columns are missing.

        Notes:
            - This method requires at least three valid observations and two unique values in the dependent variable to perform analysis.
            - Outliers are flagged based on their influence on the model's MSE, allowing for robust detection of influential points.
            - Imputation options provide flexibility for handling flagged outliers, while preserving data integrity.

        Example:
            >>> updated_df, has_outliers = DataManipulation.outlier_loo(
                    df=data,
                    parameter='dependent_var',
                    predictor='independent_var',
                    include_columns=['additional_column'],
                    mse_threshold_ratio=0.4,
                    impute='forward',
                    print_outliers=True
                )
        """
        
        # Filter valid data based on all columns in include_columns
        columns_to_check = [parameter, predictor] + include_columns
        if not isinstance(columns_to_check, list):
            raise ValueError("Parameter, predictor, and required columns must be a list of column names.")
        valid_data = df[columns_to_check].dropna().reset_index()
        x = valid_data[predictor].values
        y = valid_data[parameter].values

        modified = False  # Initialize modified flag

        if len(valid_data) > 3 and len(valid_data[parameter].unique()) > 2:
        
            # Add constant for OLS regression
            X = sm.add_constant(x)

            # Fit full model
            full_model = sm.OLS(y, X).fit()
            full_mse = np.mean((y - full_model.predict(X))**2)

            # Leave-one-out analysis
            mse_values = []
            for i in range(len(y)):
                X_loo = np.delete(X, i, axis=0)
                y_loo = np.delete(y, i)
                loo_model = sm.OLS(y_loo, X_loo).fit()
                loo_mse = np.mean((y_loo - loo_model.predict(X_loo))**2)
                mse_values.append(loo_mse)

            # Compare MSEs and flag outliers
            valid_data['MSE'] = mse_values
            valid_data['outlier'] = valid_data['MSE'].apply(
                lambda mse: '*' if mse < mse_threshold_ratio * full_mse else ''
            )

            outliers = valid_data[valid_data['outlier'] == '*']        
            
            # Optionally print outliers
            if not outliers.empty:
                modified = True  # Mark as modified
                if print_outliers:
                    print("\nOutliers detected based on OLS leave-one-out regression:")
                    print(tabulate(valid_data, headers='keys'))

            # Handle outliers
            if valid_data['outlier'].any():
                if impute == 'forward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].ffill()
                elif impute == 'backward':
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                    valid_data[parameter] = valid_data[parameter].bfill()
                elif impute is None:
                    valid_data.loc[valid_data['outlier'] == '*', parameter] = np.nan
                else:
                    raise ValueError("Invalid impute method. Use 'forward', 'backward', or None.")
                modified = True

                if print_outliers:
                    print(tabulate(valid_data, headers='keys'))

            # Replace outliers in the original DataFrame
            valid_data = valid_data.set_index('index')
            df.update(valid_data)

        return df, modified

    @staticmethod
    def mapping(df, parameter):
        """
        Create or update a mapping file for unique values in a specified column.

        Parameters:
        - df (pd.DataFrame): The input DataFrame containing the column to map.
        - parameter (str): The column name for which mappings are created or updated.

        Returns:
        - pd.DataFrame: Updated DataFrame with the specified column mapped to its corresponding values.
        """
        
        map_dir = 'output/map'
        os.makedirs(map_dir, exist_ok=True)
        map_df_name = f'map_{parameter}'
        map_path = os.path.join(map_dir, f'{map_df_name}.csv')

        if os.path.exists(map_path):
            map_df = pd.read_csv(map_path)
            existing_mapped_values = set(map_df[parameter])
            unique_values = set(df[parameter].dropna().unique())
            
            new_values = unique_values - existing_mapped_values
            if new_values:
                print(f"New unique values found for {parameter}: {new_values}")
                if len(new_values) > 10:
                    print(f"There are more than 10 new unique values for {parameter}. \nPlease fill in the new values in the CSV file and rerun the function.")
                    new_map_df = pd.DataFrame({parameter: list(new_values), 'parameter_map': [None] * len(new_values)})
                    new_map_df.to_csv(map_path, mode='a', header=False, index=False)
                    return
                else:
                    new_mappings = []
                    for value in new_values:
                        input_value = input(f"Enter the map value for new unique {value} in parameter {parameter}: ")
                        new_mappings.append({parameter: value, 'parameter_map': input_value})
                    
                    # Convert new mappings to a DataFrame and concatenate
                    new_map_df = pd.DataFrame(new_mappings)
                    map_df = pd.concat([map_df, new_map_df], ignore_index=True)
                    map_df.to_csv(map_path, index=False)
        else:
            print(f"Mapping file '{map_path}' not found. A new map will be created.")
            map_df = pd.DataFrame(columns=[parameter, 'parameter_map'])
            unique = df[parameter].dropna().unique()
            
            if len(unique) > 10:
                print(f"There are more than 10 unique values for the {parameter}. \nPlease fill in the values in the CSV file and rerun the function.")
                map_df = pd.concat([map_df, pd.DataFrame({parameter: unique, 'parameter_map': [None] * len(unique)})], ignore_index=True)
                map_df.to_csv(map_path, index=False)
                return
            else:
                mapping_dict = {}
                for i, uniq in enumerate(unique, start=1):
                    input_value = input(f"Enter the map value for {uniq} in parameter {parameter} ({i}/{len(unique)}): ")
                    if input_value.strip() == "":
                        input_value = None
                    mapping_dict[uniq] = input_value

                map_df = pd.DataFrame(list(mapping_dict.items()), columns=[parameter, 'parameter_map'])
                map_df.to_csv(map_path, index=False)

        mapping_dict = map_df.set_index(parameter)['parameter_map'].to_dict()
        df[parameter] = df[parameter].map(mapping_dict)
        return df

    @staticmethod
    def fill_missing(df, parameter, reference):
        """
        Fill missing values in a column based on reference columns and mapping files.

        Parameters:
        - df (pd.DataFrame): The input DataFrame containing the data to process.
        - parameter (str): The column name to fill missing values in.
        - reference (list): A list of reference column names used for mapping.

        Returns:
        - pd.DataFrame: Updated DataFrame with missing values in the `parameter` column filled based on reference mappings.
        """
        
        map_dir = 'output/map'
        os.makedirs(map_dir, exist_ok=True)
        null = df[parameter].isna().sum()

        for ref in reference:
            map_df_name = f'map_{ref}-{parameter}'
            map_path = os.path.join(map_dir, f'{map_df_name}.csv')

            if os.path.exists(map_path):
                map_df = pd.read_csv(map_path)
            else:
                print(f"Mapping file '{map_path}' not found. A new map will be created.")
                map_df = pd.DataFrame(columns=[ref, parameter])

            mapping_dict = map_df.set_index(ref)[parameter].to_dict()

            df[parameter] = df[parameter].fillna(df[ref].map(mapping_dict))

            null_parameters = df[df[parameter].isna()][ref].unique()
            null_parameters = [param for param in null_parameters if pd.notna(param)]

            if len(null_parameters) > 10:
                print("There are more than 10 missing reference parameters. Please fill in the values in the CSV file.")
                null_df = pd.DataFrame(null_parameters, columns=[ref])
                null_df[parameter] = None
                
                if not map_df.empty:
                    null_df = pd.concat([map_df, null_df]).drop_duplicates(subset=[ref]).reset_index(drop=True)
                
                null_df_path = os.path.join(map_dir, f'map_{ref}-{parameter}.csv')
                null_df.to_csv(null_df_path, index=False)
                print(f"Please fill in the missing values in {null_df_path} and rerun the function.")
                return
            
            elif len(null_parameters) > 0:
                for i, null_param in enumerate(null_parameters, start=1):
                    input_value = input(f"Enter the value for {null_param} in parameter {ref} ({i}/{len(null_parameters)}): ")
                    if input_value.strip() == "":
                        input_value = None

                    # Update the mapping dictionary
                    mapping_dict[null_param] = input_value

                # Re-apply the updated mapping to the DataFrame
                df[parameter] = df[parameter].fillna(df[ref].map(mapping_dict))
            
            
            # Export the updated mapping dictionary to the same map_path
            updated_map_df = pd.DataFrame(list(mapping_dict.items()), columns=[ref, parameter])
            updated_map_df.to_csv(map_path, index=False)

        null = df[parameter].isna().sum()
        percent_null = null / len(df) * 100
        print(f'Missing values in {parameter}: {null} ({percent_null:.2f}%)')

        return df

                